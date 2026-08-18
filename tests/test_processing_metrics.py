import pathlib
import tempfile
import unittest

from openpyxl import load_workbook

from scripts import processing_metrics


class ProcessingMetricsTests(unittest.TestCase):
    def test_summarizes_fresh_visual_and_all_text_usage(self):
        visual = {
            "groups": [
                {
                    "usage": {
                        "prompt_tokens": 10,
                        "completion_tokens": 4,
                        "total_tokens": 14,
                    }
                },
                {
                    "usage": {
                        "prompt_tokens": 100,
                        "completion_tokens": 50,
                        "total_tokens": 150,
                    },
                    "resumed": True,
                },
                {"usage": {"input_tokens": 3, "output_tokens": 2}},
            ]
        }
        audit = {
            "usage": {
                "initial": {
                    "prompt_tokens": 7,
                    "completion_tokens": 5,
                    "total_tokens": 12,
                },
                "repairs": [{"input_tokens": 2, "output_tokens": 1}],
            }
        }

        summary = processing_metrics.summarize_token_usage(visual, audit)

        self.assertEqual(
            summary,
            {
                "visual": {"input_tokens": 13, "output_tokens": 6, "total_tokens": 19},
                "text": {"input_tokens": 9, "output_tokens": 6, "total_tokens": 15},
                "qwen_total_tokens": 34,
                "total_tokens": 34,
            },
        )

    def test_transcription_duration_uses_declared_or_sentence_duration(self):
        transcription = {
            "transcripts": [
                {
                    "content_duration_in_milliseconds": 12_345,
                    "sentences": [{"end_time": 9_000}],
                },
                {"sentences": [{"end_time": 15_500}]},
            ]
        }
        self.assertEqual(
            processing_metrics.transcription_duration_seconds(transcription),
            15.5,
        )

    def test_builds_row_with_stage_timings_and_usage(self):
        report = {
            "video": "D:/courses/a.mp4",
            "status": "complete",
            "models": {"visual": "vl", "asr": "asr", "text": "text"},
            "stages": [
                {"name": "adaptive-keyframes", "elapsed_seconds": 1.25},
                {"name": "visual-analysis", "elapsed_seconds": 2.5},
            ],
        }
        usage = {
            "visual": {"input_tokens": 10, "output_tokens": 4, "total_tokens": 14},
            "text": {"input_tokens": 7, "output_tokens": 5, "total_tokens": 12},
            "qwen_total_tokens": 26,
        }

        row = processing_metrics.build_processing_row(
            report,
            usage,
            asr_duration_seconds=9.5,
            started_at="2026-08-14T10:00:00+08:00",
            finished_at="2026-08-14T10:00:10+08:00",
            total_elapsed_seconds=10.0,
            report_path=pathlib.Path("D:/outputs/a_report.json"),
        )

        self.assertEqual(row["视频路径"], "D:/courses/a.mp4")
        self.assertEqual(row["总耗时（秒）"], 10.0)
        self.assertEqual(row["抽帧耗时（秒）"], 1.25)
        self.assertEqual(row["视觉分析耗时（秒）"], 2.5)
        self.assertEqual(row["单视频总Token"], 26)
        self.assertEqual(row["ASR Token计量"], "不适用（按音频时长计费）")
        self.assertEqual(row["ASR音频时长（秒）"], 9.5)

    def test_xlsx_has_detail_model_and_batch_summaries_and_upserts_by_video_path(self):
        with tempfile.TemporaryDirectory() as raw:
            output = pathlib.Path(raw) / "视频处理统计.xlsx"
            base = {header: "" for header in processing_metrics.XLSX_HEADERS}
            first = {
                **base,
                "视频路径": "D:/a.mp4",
                "状态": "complete",
                "总耗时（秒）": 10,
                "视觉模型": "qwen-vl",
                "视觉总Token": 10,
                "文本模型": "qwen-text",
                "文本总Token": 4,
                "ASR模型": "paraformer-v2",
                "ASR音频时长（秒）": 12.5,
                "单视频总Token": 14,
            }
            second = {
                **base,
                "视频路径": "D:/b.mp4",
                "状态": "complete",
                "总耗时（秒）": 20,
                "视觉模型": "qwen-vl",
                "视觉总Token": 20,
                "文本模型": "qwen-text",
                "文本总Token": 6,
                "ASR模型": "paraformer-v2",
                "ASR音频时长（秒）": 15,
                "单视频总Token": 26,
            }
            updated = {
                **first,
                "视觉总Token": 30,
                "文本总Token": 10,
                "单视频总Token": 40,
            }

            processing_metrics.upsert_processing_stats(output, first)
            processing_metrics.upsert_processing_stats(output, second)
            processing_metrics.upsert_processing_stats(output, updated)

            workbook = load_workbook(output, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ["逐视频统计", "模型Token汇总", "批次汇总"],
            )
            detail = workbook["逐视频统计"]
            headers = [cell.value for cell in detail[1]]
            rows = [dict(zip(headers, values)) for values in detail.iter_rows(min_row=2, values_only=True)]
            self.assertEqual(len(rows), 2)
            by_video = {row["视频路径"]: row for row in rows}
            self.assertEqual(by_video["D:/a.mp4"]["单视频总Token"], 40)
            self.assertEqual(by_video["D:/b.mp4"]["单视频总Token"], 26)

            batch = workbook["批次汇总"]
            batch_values = dict(zip([cell.value for cell in batch[1]], [cell.value for cell in batch[2]]))
            self.assertEqual(batch_values["视频总数"], 2)
            self.assertEqual(batch_values["批次总Token"], 66)
            self.assertEqual(batch_values["视频处理总耗时（秒）"], 30)

            model_sheet = workbook["模型Token汇总"]
            model_headers = [cell.value for cell in model_sheet[1]]
            model_rows = [
                dict(zip(model_headers, values))
                for values in model_sheet.iter_rows(min_row=2, values_only=True)
            ]
            by_model = {(row["模型类型"], row["模型名称"]): row for row in model_rows}
            self.assertEqual(by_model[("视觉", "qwen-vl")]["总Token"], 50)
            self.assertEqual(by_model[("文本", "qwen-text")]["总Token"], 16)
            self.assertIsNone(by_model[("ASR", "paraformer-v2")]["总Token"])
            self.assertEqual(by_model[("ASR", "paraformer-v2")]["ASR音频时长（秒）"], 27.5)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
