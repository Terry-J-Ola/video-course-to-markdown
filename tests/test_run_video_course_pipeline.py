import contextlib
import io
import json
import os
import pathlib
import subprocess
import sys
import tempfile
import time
import unittest
import urllib.error
from unittest import mock

import yaml
from openpyxl import load_workbook

from scripts import analyze_video_frames
from scripts import assemble_evidence_markdown
from scripts import extract_adaptive_keyframes
from scripts import rewrite_learner_markdown
from scripts import run_video_course_pipeline
from scripts import transcribe_video_audio


ROOT = pathlib.Path(__file__).resolve().parents[1]
PIPELINE = ROOT / "scripts" / "run_video_course_pipeline.py"


def run_pipeline(*args: str, env: dict[str, str] | None = None) -> subprocess.CompletedProcess[str]:
    process_env = os.environ.copy()
    for name in (
        "DASHSCOPE_API_KEY",
        "DASHSCOPE_QWEN_API_KEY",
        "DASHSCOPE_ASR_API_KEY",
    ):
        process_env.pop(name, None)
    process_env["PYTHONIOENCODING"] = "utf-8"
    if env:
        process_env.update(env)
    return subprocess.run(
        [sys.executable, str(PIPELINE), *args, "--json-output"],
        cwd=ROOT,
        env=process_env,
        text=True,
        encoding="utf-8",
        capture_output=True,
        check=False,
    )


def stage(records: list[dict], name: str) -> dict:
    return next(item for item in records if item.get("stage") == name)


def option_value(command: list[str], option: str) -> str:
    return command[command.index(option) + 1]


class JsonResponse:
    def __init__(self, value: dict):
        self.payload = json.dumps(value).encode("utf-8")

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def read(self) -> bytes:
        return self.payload


def visual_response(request, timeout=240):
    payload = json.loads(request.data.decode("utf-8"))
    return JsonResponse(
        {
            "model": payload["model"],
            "output": {
                "choices": [
                    {
                        "message": {
                            "content": [{"text": json.dumps({"frames": []})}]
                        }
                    }
                ]
            },
            "usage": {},
        }
    )


def run_visual_stage(
    manifest: pathlib.Path,
    output: pathlib.Path,
    *,
    model: str = "visual-test",
    group_size: int = 4,
    min_spacing: float = 0.0,
) -> int:
    argv = [
        "analyze_video_frames.py",
        str(manifest),
        str(output),
        "--model",
        model,
        "--group-size",
        str(group_size),
        "--workers",
        "1",
        "--min-spacing",
        str(min_spacing),
    ]
    with mock.patch.object(sys, "argv", argv):
        with contextlib.redirect_stdout(io.StringIO()):
            return analyze_video_frames.main()


class DependencyPropagationTests(unittest.TestCase):
    def test_run_stage_persists_details_and_keeps_child_output_off_human_console(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            log_path = root / "视频处理日志.jsonl"
            detailed_value = str(root / "very" / "long" / "absolute" / "detail.txt")
            report = {
                "stages": [],
                "_runtime": {
                    "json_output": False,
                    "log_file": str(log_path),
                },
            }
            console = io.StringIO()

            with contextlib.redirect_stdout(console):
                run_video_course_pipeline.run_stage(
                    "visual-analysis",
                    [sys.executable, "-c", f"print({detailed_value!r})"],
                    report,
                    False,
                )

            terminal_text = console.getvalue()
            log_records = [
                json.loads(line) for line in log_path.read_text(encoding="utf-8").splitlines()
            ]
            self.assertIn("画面分析", terminal_text)
            self.assertNotIn(detailed_value, terminal_text)
            self.assertTrue(
                any(detailed_value in item.get("content", "") for item in log_records)
            )
            self.assertTrue(
                any(sys.executable in item.get("command", []) for item in log_records)
            )

    def test_run_stage_failure_persists_stderr_and_exit_code(self):
        with tempfile.TemporaryDirectory() as raw:
            log_path = pathlib.Path(raw) / "视频处理日志.jsonl"
            report = {
                "stages": [],
                "_runtime": {
                    "json_output": False,
                    "log_file": str(log_path),
                },
            }
            console = io.StringIO()
            command = [
                sys.executable,
                "-c",
                "import sys; print('FAIL-DETAIL', file=sys.stderr); raise SystemExit(7)",
            ]

            with contextlib.redirect_stdout(console):
                with self.assertRaisesRegex(subprocess.CalledProcessError, "exit status 7"):
                    run_video_course_pipeline.run_stage(
                        "audio-transcription", command, report, False
                    )

            records = [
                json.loads(line) for line in log_path.read_text(encoding="utf-8").splitlines()
            ]
            self.assertNotIn("FAIL-DETAIL", console.getvalue())
            self.assertTrue(
                any("FAIL-DETAIL" in item.get("content", "") for item in records)
            )
            failure = next(item for item in records if item["event"] == "stage_failed")
            self.assertEqual(failure["returncode"], 7)

    def test_run_stage_promotes_structured_dashscope_error(self):
        report = {"stages": []}
        detail = {
            "event": "dashscope_error",
            "service": "visual",
            "category": "authentication",
            "http_status": 401,
            "provider_code": "InvalidApiKey",
            "message": "DashScope API Key 无效或已过期",
            "retryable": False,
        }
        command = [
            sys.executable,
            "-c",
            "import json,sys; print(json.dumps(" + repr(detail) + ", ensure_ascii=False), "
            "file=sys.stderr); raise SystemExit(1)",
        ]

        with contextlib.redirect_stdout(io.StringIO()):
            with self.assertRaises(run_video_course_pipeline.StageExecutionError) as caught:
                run_video_course_pipeline.run_stage("visual-analysis", command, report, False)

        self.assertEqual(caught.exception.error_detail["category"], "authentication")
        self.assertEqual(caught.exception.error_detail["http_status"], 401)

    def test_persistent_log_redacts_unified_api_key_from_child_output(self):
        with tempfile.TemporaryDirectory() as raw:
            log_path = pathlib.Path(raw) / "视频处理日志.jsonl"
            secret = "sk-unified-must-be-redacted"
            report = {
                "stages": [],
                "_runtime": {
                    "json_output": False,
                    "log_file": str(log_path),
                },
            }
            command = [
                sys.executable,
                "-c",
                "import os; print(os.environ['DASHSCOPE_API_KEY'])",
            ]

            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": secret}, clear=False):
                with contextlib.redirect_stdout(io.StringIO()):
                    run_video_course_pipeline.run_stage(
                        "visual-analysis", command, report, False
                    )

            log_text = log_path.read_text(encoding="utf-8")
            self.assertNotIn(secret, log_text)
            self.assertIn("[REDACTED]", log_text)

    def test_streamed_command_delivers_output_before_process_completion(self):
        observed: list[tuple[str, str, float]] = []
        started = time.monotonic()
        returncode = run_video_course_pipeline.run_streamed_command(
            [
                sys.executable,
                "-c",
                "import time; print('first', flush=True); time.sleep(0.2); print('second')",
            ],
            run_video_course_pipeline.utf8_child_environment(),
            lambda stream, line: observed.append((stream, line, time.monotonic())),
        )
        finished = time.monotonic()

        self.assertEqual(returncode, 0)
        self.assertEqual([(item[0], item[1]) for item in observed], [
            ("stdout", "first"),
            ("stdout", "second"),
        ])
        self.assertLess(observed[0][2] - started, finished - observed[0][2])

    def test_persistent_log_rotates_and_preserves_run_id(self):
        with tempfile.TemporaryDirectory() as raw:
            log_path = pathlib.Path(raw) / "视频处理日志.jsonl"
            with mock.patch.object(run_video_course_pipeline, "ACTIVE_RUN_ID", "run-test"):
                with mock.patch.object(run_video_course_pipeline, "LOG_MAX_BYTES", 240):
                    with mock.patch.object(run_video_course_pipeline, "LOG_BACKUPS", 2):
                        for index in range(5):
                            run_video_course_pipeline.append_log(
                                log_path,
                                {"event": "probe", "index": index, "content": "x" * 100},
                            )

            self.assertTrue(log_path.is_file())
            self.assertTrue(log_path.with_name(f"{log_path.name}.1").is_file())
            records = []
            for candidate in log_path.parent.glob(f"{log_path.name}*"):
                records.extend(
                    json.loads(line)
                    for line in candidate.read_text(encoding="utf-8").splitlines()
                )
            self.assertTrue(records)
            self.assertTrue(all(item["run_id"] == "run-test" for item in records))

    def test_child_python_processes_force_utf8_streams(self):
        report = {"stages": []}
        completed = subprocess.CompletedProcess(["python"], 0)
        with mock.patch.dict(
            os.environ,
            {"PYTHONIOENCODING": "legacy", "PYTHONUTF8": "0"},
            clear=False,
        ):
            with mock.patch.object(
                run_video_course_pipeline,
                "run_streamed_command",
                return_value=0,
            ) as stage_runner:
                with mock.patch.object(
                    run_video_course_pipeline.subprocess,
                    "run",
                    return_value=completed,
                ) as batch_runner:
                    run_video_course_pipeline.run_stage(
                        "utf8-stage", ["python", "stage.py"], report, False
                    )
                    self.assertEqual(
                        run_video_course_pipeline.run_batch_child(["python", "batch.py"]), 0
                    )

        stage_environment = stage_runner.call_args.args[1]
        batch_environment = batch_runner.call_args.kwargs["env"]
        for environment in (stage_environment, batch_environment):
            self.assertEqual(environment["PYTHONIOENCODING"], "utf-8")
            self.assertEqual(environment["PYTHONUTF8"], "1")

    def test_pydeps_are_prepended_for_real_extraction_and_contact_sheet_imports(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            pydeps = root / "pydeps"
            pil = pydeps / "PIL"
            pil.mkdir(parents=True)
            (pydeps / "numpy.py").write_text(
                "class ndarray: pass\nfloat32 = object()\nuint8 = object()\n",
                encoding="utf-8",
            )
            (pil / "__init__.py").write_text(
                "Image = object()\nImageFilter = object()\n"
                "ImageDraw = object()\nImageFont = object()\n",
                encoding="utf-8",
            )
            inherited = root / "inherited"
            inherited.mkdir()
            (inherited / "inherited_probe.py").write_text("VALUE = 'preserved'\n", encoding="utf-8")
            secret = "sk-child-env-must-not-enter-report"
            report = {"pydeps": str(pydeps), "stages": []}

            commands = [
                [sys.executable, "-S", str(ROOT / "scripts" / "extract_adaptive_keyframes.py"), "--help"],
                [sys.executable, "-S", str(ROOT / "scripts" / "make_contact_sheet.py"), "--help"],
                [
                    sys.executable,
                    "-S",
                    "-c",
                    "import numpy, inherited_probe; assert inherited_probe.VALUE == 'preserved'",
                ],
            ]
            with mock.patch.dict(
                os.environ,
                {"PYTHONPATH": str(inherited), "DASHSCOPE_API_KEY": secret},
                clear=False,
            ):
                for index, command in enumerate(commands):
                    with self.subTest(command=index):
                        run_video_course_pipeline.run_stage(
                            f"dependency-probe-{index}", command, report, False
                        )

            self.assertNotIn(secret, json.dumps(report))


class ResumeProvenanceTests(unittest.TestCase):
    def make_visual_manifest(self, root: pathlib.Path, frame_count: int = 1):
        video = root / "course.mp4"
        video.write_bytes(b"video-v1")
        keyframes = []
        for index in range(frame_count):
            frame = root / f"frame-{index}.jpg"
            frame.write_bytes(f"frame-{index}".encode("ascii"))
            keyframes.append(
                {"timestamp_seconds": float(index), "image": str(frame)}
            )
        manifest = root / "manifest.json"
        manifest.write_text(
            json.dumps({"video": str(video), "keyframes": keyframes}),
            encoding="utf-8",
        )
        return video, manifest

    def test_visual_json_parser_recovers_valid_object_from_wrapped_prose(self):
        content = (
            '说明中的示例对象 {"status":"illustration"}；实际结果如下：'
            '{"frames":[{"frame_index":1}]}，以上。'
        )

        parsed = analyze_video_frames.parse_json_content(content)

        self.assertEqual(parsed, {"frames": [{"frame_index": 1}]})

    def test_visual_call_uses_dashscope_native_multimodal_protocol(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames, "open_url", side_effect=visual_response
                ) as urlopen:
                    self.assertEqual(run_visual_stage(manifest, output), 0)

            request = urlopen.call_args.args[0]
            payload = json.loads(request.data.decode("utf-8"))
            self.assertEqual(
                request.full_url,
                "https://dashscope.aliyuncs.com/api/v1/services/aigc/"
                "multimodal-generation/generation",
            )
            self.assertNotIn("messages", payload)
            self.assertIn("messages", payload["input"])
            content = payload["input"]["messages"][0]["content"]
            self.assertTrue(any("text" in item for item in content))
            self.assertTrue(any(item.get("image", "").startswith("data:image/") for item in content))
            self.assertEqual(payload["parameters"]["result_format"], "message")
            self.assertEqual(
                payload["parameters"]["response_format"], {"type": "json_object"}
            )
            checkpoint = json.loads(
                (output / "group_0001_checkpoint.json").read_text(encoding="utf-8")
            )
            self.assertEqual(
                checkpoint["fingerprint"]["api_protocol"],
                "dashscope-native-multimodal-v1",
            )

    def test_visual_authentication_error_is_not_retried(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            error = urllib.error.HTTPError(
                "https://dashscope.aliyuncs.com/api",
                401,
                "unauthorized",
                {},
                io.BytesIO(b'{"code":"InvalidApiKey","message":"invalid key"}'),
            )
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(analyze_video_frames.time, "sleep") as sleep:
                    with mock.patch.object(
                        analyze_video_frames, "open_url", side_effect=error
                    ) as urlopen:
                        with self.assertRaisesRegex(RuntimeError, "API Key"):
                            run_visual_stage(manifest, output)

            self.assertEqual(urlopen.call_count, 1)
            sleep.assert_not_called()

    def test_visual_rate_limit_retries_and_honors_retry_after(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            limited = urllib.error.HTTPError(
                "https://dashscope.aliyuncs.com/api",
                429,
                "limited",
                {"Retry-After": "1.5"},
                io.BytesIO(b'{"code":"Throttling","message":"slow down"}'),
            )
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(analyze_video_frames.time, "sleep") as sleep:
                    with mock.patch.object(
                        analyze_video_frames,
                        "open_url",
                        side_effect=[limited, visual_response(mock.Mock(data=b'{"model":"visual-test"}'))],
                    ) as urlopen:
                        self.assertEqual(run_visual_stage(manifest, output), 0)

            self.assertEqual(urlopen.call_count, 2)
            sleep.assert_called_once_with(1.5)

    def test_visual_checkpoint_is_reused_when_provenance_is_compatible(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames, "open_url", side_effect=visual_response
                ) as urlopen:
                    self.assertEqual(run_visual_stage(manifest, output), 0)
                    self.assertEqual(run_visual_stage(manifest, output), 0)
            self.assertEqual(urlopen.call_count, 1)
            consolidated = json.loads((output / "consolidated.json").read_text(encoding="utf-8"))
            self.assertTrue(consolidated["groups"][0]["resumed"])

    def test_visual_checkpoint_is_rejected_when_model_changes(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames, "open_url", side_effect=visual_response
                ) as urlopen:
                    self.assertEqual(run_visual_stage(manifest, output, model="visual-old"), 0)
                    self.assertEqual(run_visual_stage(manifest, output, model="visual-new"), 0)
            self.assertEqual(urlopen.call_count, 2)
            consolidated = json.loads((output / "consolidated.json").read_text(encoding="utf-8"))
            self.assertEqual(consolidated["model"], "visual-new")
            self.assertNotIn("resumed", consolidated["groups"][0])

    def test_visual_checkpoint_is_rejected_when_group_parameters_change(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root, frame_count=2)
            output = root / "visual"
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames, "open_url", side_effect=visual_response
                ) as urlopen:
                    self.assertEqual(run_visual_stage(manifest, output, group_size=2), 0)
                    self.assertEqual(run_visual_stage(manifest, output, group_size=1), 0)
            self.assertEqual(urlopen.call_count, 3)

    def test_visual_checkpoint_is_rejected_when_min_spacing_changes(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root, frame_count=2)
            output = root / "visual"
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames, "open_url", side_effect=visual_response
                ) as urlopen:
                    self.assertEqual(run_visual_stage(manifest, output), 0)
                    self.assertEqual(run_visual_stage(manifest, output, min_spacing=0.5), 0)
            self.assertEqual(urlopen.call_count, 2)

    def test_visual_checkpoint_is_rejected_when_input_identity_changes(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames, "open_url", side_effect=visual_response
                ) as urlopen:
                    self.assertEqual(run_visual_stage(manifest, output), 0)
                    video.write_bytes(b"video-version-two-is-different")
                    self.assertEqual(run_visual_stage(manifest, output), 0)
            self.assertEqual(urlopen.call_count, 2)

    def test_visual_consolidated_metadata_uses_the_artifact_producer_model(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"

            def provider_response(request, timeout=240):
                return JsonResponse(
                    {
                        "model": "provider-resolved-visual-model",
                        "choices": [
                            {"message": {"content": json.dumps({"frames": []})}}
                        ],
                        "usage": {},
                    }
                )

            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames,
                    "open_url",
                    side_effect=provider_response,
                ):
                    self.assertEqual(
                        run_visual_stage(manifest, output, model="requested-visual-alias"), 0
                    )
            consolidated = json.loads((output / "consolidated.json").read_text(encoding="utf-8"))
            self.assertEqual(consolidated["model"], "provider-resolved-visual-model")

    def test_invalid_visual_response_does_not_commit_a_resumable_checkpoint(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            invalid = JsonResponse(
                {
                    "model": "visual-test",
                    "choices": [{"message": {"content": "not-json"}}],
                    "usage": {},
                }
            )
            valid = JsonResponse(
                {
                    "model": "visual-test",
                    "choices": [
                        {"message": {"content": json.dumps({"frames": []})}}
                    ],
                    "usage": {},
                }
            )
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames.time, "sleep", return_value=None
                ):
                    diagnostics = io.StringIO()
                    with contextlib.redirect_stderr(diagnostics):
                        with mock.patch.object(
                            analyze_video_frames,
                            "open_url",
                            side_effect=[invalid, invalid, invalid],
                        ):
                            with self.assertRaisesRegex(
                                RuntimeError, "group 1 failed after 3 attempts"
                            ):
                                run_visual_stage(manifest, output)

                    checkpoint = output / "group_0001_checkpoint.json"
                    self.assertFalse(checkpoint.exists())
                    raw_response = json.loads(
                        (output / "group_0001_raw.json").read_text(encoding="utf-8")
                    )
                    self.assertEqual(raw_response["choices"][0]["message"]["content"], "not-json")
                    diagnostic_records = [
                        json.loads(line) for line in diagnostics.getvalue().splitlines()
                    ]
                    self.assertEqual([item["attempt"] for item in diagnostic_records], [1, 2, 3])
                    self.assertTrue(
                        all(item["error_type"] == "JSONDecodeError" for item in diagnostic_records)
                    )
                    with mock.patch.object(
                        analyze_video_frames,
                        "open_url",
                        return_value=valid,
                    ) as urlopen:
                        self.assertEqual(run_visual_stage(manifest, output), 0)
            urlopen.assert_called_once()

    def test_structurally_malformed_visual_response_retries_without_checkpoint(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            malformed = JsonResponse({"model": "visual-test", "choices": []})
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames.time, "sleep", return_value=None
                ):
                    with mock.patch.object(
                        analyze_video_frames,
                        "open_url",
                        side_effect=[malformed, malformed, malformed],
                    ) as urlopen:
                        with self.assertRaisesRegex(RuntimeError, "group 1 failed"):
                            run_visual_stage(manifest, output)
            self.assertEqual(urlopen.call_count, 3)
            self.assertFalse((output / "group_0001_checkpoint.json").exists())

    def test_schema_invalid_visual_result_retries_without_checkpoint(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            responses = [
                JsonResponse(
                    {
                        "model": "visual-test",
                        "choices": [{"message": {"content": payload}}],
                    }
                )
                for payload in ("[]", "{}", '{"frames":"not-a-list"}')
            ]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames.time, "sleep", return_value=None
                ):
                    with mock.patch.object(
                        analyze_video_frames,
                        "open_url",
                        side_effect=responses,
                    ) as urlopen:
                        with self.assertRaisesRegex(RuntimeError, "group 1 failed"):
                            run_visual_stage(manifest, output)
            self.assertEqual(urlopen.call_count, 3)
            self.assertFalse((output / "group_0001_checkpoint.json").exists())

    def test_non_object_visual_envelope_retries_without_checkpoint(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            responses = [
                JsonResponse([]),
                JsonResponse({"choices": [None]}),
                JsonResponse({"choices": ["not-an-object"]}),
            ]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(analyze_video_frames.time, "sleep", return_value=None):
                    with mock.patch.object(
                        analyze_video_frames,
                        "open_url",
                        side_effect=responses,
                    ) as urlopen:
                        with self.assertRaisesRegex(RuntimeError, "group 1 failed"):
                            run_visual_stage(manifest, output)
            self.assertEqual(urlopen.call_count, 3)
            self.assertFalse((output / "group_0001_checkpoint.json").exists())

    def test_non_object_resumed_visual_envelope_invalidates_and_recomputes(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames, "open_url", side_effect=visual_response
                ) as urlopen:
                    self.assertEqual(run_visual_stage(manifest, output), 0)
                    (output / "group_0001_raw.json").write_text("[]\n", encoding="utf-8")
                    self.assertEqual(run_visual_stage(manifest, output), 0)
            self.assertEqual(urlopen.call_count, 2)

    def test_corrupt_visual_checkpoint_artifact_is_recomputed(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root)
            output = root / "visual"
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames, "open_url", side_effect=visual_response
                ) as urlopen:
                    self.assertEqual(run_visual_stage(manifest, output), 0)
                    (output / "group_0001_raw.json").write_text(
                        json.dumps(
                            {
                                "model": "visual-test",
                                "choices": [{"message": {"content": "not-json"}}],
                                "usage": {},
                            }
                        ),
                        encoding="utf-8",
                    )
                    self.assertEqual(run_visual_stage(manifest, output), 0)
            self.assertEqual(urlopen.call_count, 2)

    def test_mixed_visual_producers_are_reported_without_requested_alias(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            _, manifest = self.make_visual_manifest(root, frame_count=2)
            output = root / "visual"

            def mixed_response(request, timeout=240):
                payload = json.loads(request.data.decode("utf-8"))
                current = mixed_response.calls
                mixed_response.calls += 1
                return JsonResponse(
                    {
                        "model": f"provider-model-{current + 1}",
                        "choices": [
                            {"message": {"content": json.dumps({"frames": []})}}
                        ],
                        "usage": {},
                    }
                )

            mixed_response.calls = 0
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    analyze_video_frames,
                    "open_url",
                    side_effect=mixed_response,
                ):
                    self.assertEqual(
                        run_visual_stage(
                            manifest,
                            output,
                            model="requested-visual-alias",
                            group_size=1,
                        ),
                        0,
                    )
            consolidated = json.loads((output / "consolidated.json").read_text(encoding="utf-8"))
            self.assertEqual(consolidated["model"], "mixed")
            self.assertEqual(
                consolidated["producer_models"],
                ["provider-model-1", "provider-model-2"],
            )
            self.assertNotEqual(consolidated["model"], "requested-visual-alias")

    def test_skip_asr_checkpoint_does_not_resume_as_a_normal_transcription(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"video")
            output = root / "output"

            def complete_stage(name: str, command: list[str], report: dict, dry_run: bool) -> None:
                report["stages"].append({"name": name, "command": command, "status": "complete"})
                if name == "adaptive-keyframes":
                    frames_dir = pathlib.Path(command[3])
                    frames_dir.mkdir(parents=True, exist_ok=True)
                    (frames_dir / "manifest.json").write_text(
                        json.dumps({"mode": "slides"}), encoding="utf-8"
                    )
                elif name == "visual-analysis":
                    visual_dir = pathlib.Path(command[3])
                    visual_dir.mkdir(parents=True, exist_ok=True)
                    (visual_dir / "consolidated.json").write_text(
                        json.dumps({"model": option_value(command, "--model"), "groups": []}),
                        encoding="utf-8",
                    )

            pipeline_argv = [
                str(PIPELINE),
                str(video),
                str(output),
                "--skip-asr",
                "--skip-business",
            ]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", pipeline_argv):
                    with mock.patch.object(run_video_course_pipeline, "run_stage", complete_stage):
                        with contextlib.redirect_stdout(io.StringIO()):
                            self.assertEqual(run_video_course_pipeline.main(), 0)

            asr_dir = output / "_video_course_work" / "course" / "asr"

            class FakeOssUtils:
                @staticmethod
                def upload(model, file_path, api_key):
                    return "oss://test/audio", None

            def fake_extract_audio(ffmpeg, source_video, audio):
                audio.parent.mkdir(parents=True, exist_ok=True)
                audio.write_bytes(b"audio")

            def fake_request_json(url, api_key, method="GET", payload=None, oss_resolve=False):
                if method == "POST":
                    return {"output": {"task_id": "task-1"}}
                return {
                    "output": {
                        "task_status": "SUCCEEDED",
                        "result": {"transcription_url": "https://example.invalid/result"},
                    }
                }

            transcript = {
                "transcripts": [
                    {"sentences": [{"text": "fresh", "begin_time": 0, "end_time": 1000}]}
                ]
            }
            transcribe_argv = [
                "transcribe_video_audio.py",
                str(video),
                str(asr_dir),
                "--model",
                "paraformer-v2",
            ]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", transcribe_argv):
                    with mock.patch.object(
                        transcribe_video_audio,
                        "configure_dependencies",
                        return_value=("ffmpeg", FakeOssUtils),
                    ) as configure:
                        with mock.patch.object(
                            transcribe_video_audio, "extract_audio", side_effect=fake_extract_audio
                        ):
                            with mock.patch.object(
                                transcribe_video_audio, "request_json", side_effect=fake_request_json
                            ):
                                with mock.patch.object(
                                    transcribe_video_audio,
                                    "open_url",
                                    return_value=JsonResponse(transcript),
                                ):
                                    with contextlib.redirect_stdout(io.StringIO()):
                                        self.assertEqual(transcribe_video_audio.main(), 0)

            configure.assert_called_once()
            saved = json.loads((asr_dir / "transcription.json").read_text(encoding="utf-8"))
            self.assertEqual(saved["transcripts"][0]["sentences"][0]["text"], "fresh")

    def test_asr_checkpoint_reuse_requires_matching_input_and_model(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"video-v1")
            output = root / "asr"
            output.mkdir()
            transcription = {"transcripts": [{"sentences": []}]}
            (output / "transcription.json").write_text(
                json.dumps(transcription), encoding="utf-8"
            )
            fingerprint = {
                "stage": "asr",
                "input": transcribe_video_audio.file_identity(video),
                "state": "transcribed",
                "model": "paraformer-v2",
            }
            transcribe_video_audio.write_checkpoint(
                output / "transcription.checkpoint.json", fingerprint
            )

            def invoke(model: str):
                argv = [
                    "transcribe_video_audio.py",
                    str(video),
                    str(output),
                    "--model",
                    model,
                ]
                with mock.patch.object(sys, "argv", argv):
                    with contextlib.redirect_stdout(io.StringIO()):
                        return transcribe_video_audio.main()

            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(
                    transcribe_video_audio,
                    "configure_dependencies",
                    side_effect=RuntimeError("incompatible checkpoint reached recomputation"),
                ) as configure:
                    self.assertEqual(invoke("paraformer-v2"), 0)
                    self.assertEqual(configure.call_count, 0)
                    with self.assertRaisesRegex(RuntimeError, "incompatible checkpoint"):
                        invoke("another-asr-model")
                    video.write_bytes(b"video-version-two")
                    with self.assertRaisesRegex(RuntimeError, "incompatible checkpoint"):
                        invoke("paraformer-v2")

    def test_corrupt_asr_checkpoint_artifact_is_recomputed(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"video")
            output = root / "asr"
            output.mkdir()
            transcription = output / "transcription.json"
            transcription.write_text("not-json", encoding="utf-8")
            checkpoint = output / "transcription.checkpoint.json"
            transcribe_video_audio.write_checkpoint(
                checkpoint,
                {
                    "stage": "asr",
                    "input": transcribe_video_audio.file_identity(video),
                    "state": "transcribed",
                    "model": "paraformer-v2",
                },
            )
            argv = [
                "transcribe_video_audio.py",
                str(video),
                str(output),
                "--model",
                "paraformer-v2",
            ]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(
                        transcribe_video_audio,
                        "configure_dependencies",
                        side_effect=RuntimeError("corrupt checkpoint reached recomputation"),
                    ) as configure:
                        with self.assertRaisesRegex(RuntimeError, "recomputation"):
                            with contextlib.redirect_stdout(io.StringIO()):
                                transcribe_video_audio.main()
            configure.assert_called_once()
            self.assertFalse(checkpoint.exists())

    def test_nested_corrupt_asr_checkpoint_artifact_is_recomputed(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"video")
            output = root / "asr"
            output.mkdir()
            transcription = output / "transcription.json"
            transcription.write_text(
                json.dumps({"transcripts": [{"sentences": ["bad-entry"]}]}),
                encoding="utf-8",
            )
            checkpoint = output / "transcription.checkpoint.json"
            transcribe_video_audio.write_checkpoint(
                checkpoint,
                {
                    "stage": "asr",
                    "input": transcribe_video_audio.file_identity(video),
                    "state": "transcribed",
                    "model": "paraformer-v2",
                },
            )
            argv = [
                "transcribe_video_audio.py",
                str(video),
                str(output),
                "--model",
                "paraformer-v2",
            ]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(
                        transcribe_video_audio,
                        "configure_dependencies",
                        side_effect=RuntimeError("nested corrupt checkpoint reached recomputation"),
                    ) as configure:
                        with self.assertRaisesRegex(RuntimeError, "recomputation"):
                            with contextlib.redirect_stdout(io.StringIO()):
                                transcribe_video_audio.main()
            configure.assert_called_once()
            self.assertFalse(checkpoint.exists())

    def test_malformed_fresh_asr_result_is_not_checkpointed(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            transcription_path = root / "transcription.json"
            checkpoint_path = root / "transcription.checkpoint.json"
            malformed = {"transcripts": [{"sentences": ["bad-entry"]}]}
            with self.assertRaisesRegex(ValueError, "sentence"):
                transcribe_video_audio.persist_transcription(
                    transcription_path,
                    checkpoint_path,
                    malformed,
                    {"stage": "asr"},
                )
            self.assertFalse(transcription_path.exists())
            self.assertFalse(checkpoint_path.exists())

    def test_extracted_audio_reuse_requires_matching_input_identity(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"video-v1")
            audio = root / "audio.mp3"
            audio.write_bytes(b"audio-from-v1")
            checkpoint = root / "audio.checkpoint.json"
            fingerprint = transcribe_video_audio.audio_fingerprint(video)
            transcribe_video_audio.write_checkpoint(checkpoint, fingerprint)

            calls = []

            def fake_extract(ffmpeg, source_video, target):
                self.assertFalse(target.exists())
                target.write_bytes(b"audio-from-v2")
                calls.append(source_video.read_bytes())

            with mock.patch.object(
                transcribe_video_audio, "extract_audio", side_effect=fake_extract
            ):
                transcribe_video_audio.ensure_audio("ffmpeg", video, audio, checkpoint)
                self.assertEqual(calls, [])
                video.write_bytes(b"video-version-two")
                transcribe_video_audio.ensure_audio("ffmpeg", video, audio, checkpoint)

            self.assertEqual(calls, [b"video-version-two"])
            self.assertEqual(audio.read_bytes(), b"audio-from-v2")

    def test_keyframe_scan_reuse_requires_matching_input_and_scan_parameters(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"video-v1")
            output = root / "frames"
            scan_dir = output / "scan"
            scan_dir.mkdir(parents=True)
            (scan_dir / "scan_000001.jpg").write_bytes(b"old-frame")
            fingerprint = {
                "stage": "adaptive-keyframes",
                "input": extract_adaptive_keyframes.file_identity(video),
                "mode": "slides",
                "fps": 2.0,
            }
            extract_adaptive_keyframes.write_checkpoint(
                output / "extraction.checkpoint.json", fingerprint
            )

            self.assertTrue(
                extract_adaptive_keyframes.scan_checkpoint_is_compatible(
                    video, output, "slides", 2.0
                )
            )
            self.assertFalse(
                extract_adaptive_keyframes.scan_checkpoint_is_compatible(
                    video, output, "slides", 3.0
                )
            )
            video.write_bytes(b"video-version-two")
            self.assertFalse(
                extract_adaptive_keyframes.scan_checkpoint_is_compatible(
                    video, output, "slides", 2.0
                )
            )


class PipelineCliTests(unittest.TestCase):
    def test_default_dry_run_console_hides_long_absolute_paths(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "很长的课程目录" / "更长的章节目录" / "课程.mp4"
            video.parent.mkdir(parents=True)
            video.write_bytes(b"placeholder")
            output = root / "同样很长的输出目录" / "output"
            process_env = os.environ.copy()
            process_env["PYTHONIOENCODING"] = "utf-8"

            result = subprocess.run(
                [sys.executable, str(PIPELINE), str(video), str(output), "--dry-run"],
                cwd=ROOT,
                env=process_env,
                text=True,
                encoding="utf-8",
                capture_output=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertIn("课程.mp4", result.stdout)
            self.assertIn("计划", result.stdout)
            self.assertNotIn(str(video.resolve()), result.stdout)
            self.assertNotIn(str(output.resolve()), result.stdout)
            self.assertFalse(output.exists())

    def test_pipeline_module_is_importable_from_repository_root(self):
        result = subprocess.run(
            [
                sys.executable,
                "-c",
                (
                    "from scripts.run_video_course_pipeline import "
                    "validate_output_dir, validate_video; "
                    "print(validate_output_dir.__name__, validate_video.__name__)"
                ),
            ],
            cwd=ROOT,
            text=True,
            encoding="utf-8",
            capture_output=True,
            check=False,
        )
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(result.stdout.strip(), "validate_output_dir validate_video")

    def test_dry_run_uses_independent_default_models_without_creating_output(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            output = root / "output"
            result = run_pipeline(str(video), str(output), "--dry-run")
            self.assertEqual(result.returncode, 0, result.stderr)
            records = [json.loads(line) for line in result.stdout.splitlines()]
            summary = records[-1]
            self.assertEqual(
                summary["models"],
                {"visual": "qwen3-vl-flash", "asr": "paraformer-v2", "text": "qwen-plus"},
            )
            learner = next(item for item in records if item.get("stage") == "learner-markdown-and-audit")
            self.assertIn("qwen-plus", learner["command"])
            visual = next(item for item in records if item.get("stage") == "visual-analysis")
            self.assertIn("qwen3-vl-flash", visual["command"])
            self.assertNotIn("qwen-plus", visual["command"])
            self.assertTrue(summary["outputs"]["processing_stats"].endswith("视频处理统计.xlsx"))
            self.assertFalse(output.exists())

    def test_directory_dry_run_processes_supported_videos_in_sorted_order(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "b.mp4").write_bytes(b"placeholder")
            (input_dir / "a.mov").write_bytes(b"placeholder")
            (input_dir / "ignore.txt").write_text("ignore", encoding="utf-8")
            output = root / "output"

            result = run_pipeline(str(input_dir), str(output), "--dry-run")

            self.assertEqual(result.returncode, 0, result.stderr)
            records = [json.loads(line) for line in result.stdout.splitlines()]
            batch = records[-1]
            self.assertEqual(batch["status"], "planned")
            self.assertEqual(batch["total_videos"], 2)
            self.assertEqual(batch["planned_videos"], 2)
            starts = [item for item in records if item.get("stage") == "batch-video"]
            self.assertEqual([pathlib.Path(item["video"]).name for item in starts], ["a.mov", "b.mp4"])
            self.assertTrue(batch["processing_stats"].endswith("视频处理统计.xlsx"))
            self.assertTrue(batch["business_lectures_dir"].endswith("业务讲义汇总"))
            self.assertEqual(
                [item["business_lecture"]["status"] for item in batch["results"]],
                ["planned", "planned"],
            )
            self.assertFalse(output.exists())

    def test_batch_continues_after_failure_and_writes_batch_report(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            input_dir = root / "input"
            input_dir.mkdir()
            first = input_dir / "a.mp4"
            second = input_dir / "b.mp4"
            first.write_bytes(b"placeholder")
            second.write_bytes(b"placeholder")
            output = root / "output"
            argv = [str(PIPELINE), str(input_dir), str(output)]

            with mock.patch.object(sys, "argv", argv):
                with mock.patch.object(
                    run_video_course_pipeline,
                    "run_batch_child",
                    side_effect=[0, 7],
                ) as child:
                    with contextlib.redirect_stdout(io.StringIO()):
                        self.assertEqual(run_video_course_pipeline.main(), 1)

            self.assertEqual(child.call_count, 2)
            report_path = output / "批量处理报告.json"
            report = json.loads(report_path.read_text(encoding="utf-8"))
            self.assertEqual(report["status"], "partial")
            self.assertEqual(report["completed_videos"], 1)
            self.assertEqual(report["failed_videos"], 1)
            self.assertEqual(report["business_lecture_failures"], 1)
            self.assertEqual(
                [item["status"] for item in report["results"]],
                ["complete", "failed"],
            )
            self.assertEqual(report["results"][0]["business_lecture"]["status"], "failed")
            self.assertEqual(report["results"][1]["business_lecture"]["status"], "not-created")
            stats_path = output / "视频处理统计.xlsx"
            workbook = load_workbook(stats_path, data_only=True)
            detail = workbook["逐视频统计"]
            headers = [cell.value for cell in detail[1]]
            rows = [dict(zip(headers, values)) for values in detail.iter_rows(min_row=2, values_only=True)]
            failed = next(row for row in rows if row["视频路径"] == str(second.resolve()))
            self.assertEqual(failed["状态"], "failed")
            self.assertIn("exit code 7", failed["错误信息"])
            self.assertTrue(pathlib.Path(failed["处理报告"]).is_file())
            workbook.close()

    def test_batch_copies_business_lectures_to_flat_collection_and_keeps_originals(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            input_dir = root / "input"
            first = input_dir / "章节一" / "同名课程.mp4"
            second = input_dir / "章节二" / "同名课程.mp4"
            first.parent.mkdir(parents=True)
            second.parent.mkdir(parents=True)
            first.write_bytes(b"placeholder")
            second.write_bytes(b"placeholder")
            output = root / "output"
            argv = [str(PIPELINE), str(input_dir), str(output), "--recursive"]

            def successful_child(command: list[str]) -> int:
                video = pathlib.Path(command[2])
                video_output = pathlib.Path(command[3])
                video_output.mkdir(parents=True, exist_ok=True)
                (video_output / f"{video.stem}_业务讲义.md").write_text(
                    f"# {video.parent.name}",
                    encoding="utf-8",
                )
                return 0

            console = io.StringIO()
            with mock.patch.object(sys, "argv", argv):
                with mock.patch.object(
                    run_video_course_pipeline,
                    "run_batch_child",
                    side_effect=successful_child,
                ):
                    with contextlib.redirect_stdout(console):
                        self.assertEqual(run_video_course_pipeline.main(), 0)

            original_first = output / "章节一" / "同名课程" / "同名课程_业务讲义.md"
            original_second = output / "章节二" / "同名课程" / "同名课程_业务讲义.md"
            collection = output / "业务讲义汇总"
            self.assertEqual(original_first.read_text(encoding="utf-8"), "# 章节一")
            self.assertEqual(original_second.read_text(encoding="utf-8"), "# 章节二")

            report = json.loads((output / "批量处理报告.json").read_text(encoding="utf-8"))
            collected = [
                pathlib.Path(item["business_lecture"]["destination"])
                for item in report["results"]
            ]
            self.assertEqual(
                [path.read_text(encoding="utf-8") for path in collected],
                ["# 章节一", "# 章节二"],
            )
            self.assertEqual(len({path.name.casefold() for path in collected}), 2)
            self.assertEqual(report["business_lectures_dir"], str(collection.resolve()))
            self.assertTrue(pathlib.Path(report["business_lectures_manifest"]).is_file())
            self.assertEqual(
                [item["business_lecture"]["status"] for item in report["results"]],
                ["copied", "copied"],
            )
            self.assertNotIn(str(first.resolve()), console.getvalue())
            self.assertNotIn(str(output.resolve()), console.getvalue())
            log_path = pathlib.Path(report["log_file"])
            log_records = [
                json.loads(line) for line in log_path.read_text(encoding="utf-8").splitlines()
            ]
            self.assertEqual(log_path, output / "视频处理日志.jsonl")
            self.assertTrue(
                any(item.get("video") == str(first.resolve()) for item in log_records)
            )
            self.assertTrue(
                {"batch_started", "batch_video_started", "batch_finished"}.issubset(
                    {item["event"] for item in log_records}
                )
            )
            self.assertTrue(all(item["run_id"] == report["run_id"] for item in log_records))

    def test_flat_collection_names_are_collision_and_length_safe(self):
        output_root = pathlib.Path("D:/output")
        first = run_video_course_pipeline.batch_business_lecture_paths(
            pathlib.Path("D:/input/a__b/course.mp4"),
            output_root / "a__b" / "course",
            output_root,
        )[1]
        second = run_video_course_pipeline.batch_business_lecture_paths(
            pathlib.Path("D:/input/a/b/course.mp4"),
            output_root / "a" / "b" / "course",
            output_root,
        )[1]
        long_destination = run_video_course_pipeline.batch_business_lecture_paths(
            pathlib.Path("D:/input/long/course.mp4"),
            output_root / ("很长目录" * 40) / ("很长课程" * 30),
            output_root,
        )[1]

        self.assertNotEqual(first.name.casefold(), second.name.casefold())
        self.assertLessEqual(len(long_destination.name), 220)

    def test_batch_manifest_removes_stale_managed_lecture_but_keeps_user_file(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            input_dir = root / "input"
            input_dir.mkdir()
            first = input_dir / "a.mp4"
            second = input_dir / "b.mp4"
            first.write_bytes(b"placeholder")
            second.write_bytes(b"placeholder")
            output = root / "output"

            def successful_child(command: list[str]) -> int:
                video = pathlib.Path(command[2])
                video_output = pathlib.Path(command[3])
                video_output.mkdir(parents=True, exist_ok=True)
                (video_output / f"{video.stem}_业务讲义.md").write_text(
                    video.name,
                    encoding="utf-8",
                )
                return 0

            def run_once() -> int:
                argv = [str(PIPELINE), str(input_dir), str(output)]
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(
                        run_video_course_pipeline,
                        "run_batch_child",
                        side_effect=successful_child,
                    ):
                        with contextlib.redirect_stdout(io.StringIO()):
                            return run_video_course_pipeline.main()

            self.assertEqual(run_once(), 0)
            collection = output / "业务讲义汇总"
            manifest_path = collection / "业务讲义汇总清单.json"
            first_manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            second_entry = next(
                item for item in first_manifest["entries"] if item["video"] == str(second.resolve())
            )
            stale_path = collection / second_entry["filename"]
            first_entry = next(
                item for item in first_manifest["entries"] if item["video"] == str(first.resolve())
            )
            legacy_path = collection / "legacy-managed-name.md"
            (collection / first_entry["filename"]).replace(legacy_path)
            first_entry["filename"] = legacy_path.name
            manifest_path.write_text(
                json.dumps(first_manifest, ensure_ascii=False, indent=2) + "\n",
                encoding="utf-8",
            )
            user_file = collection / "用户手工文档.md"
            user_file.write_text("keep", encoding="utf-8")

            second.unlink()
            self.assertEqual(run_once(), 0)

            self.assertFalse(stale_path.exists())
            self.assertFalse(legacy_path.exists())
            self.assertEqual(user_file.read_text(encoding="utf-8"), "keep")
            second_manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(second_manifest["schema_version"], 1)
            self.assertEqual(second_manifest["pending_cleanup"], [])
            self.assertEqual(
                [item["video"] for item in second_manifest["entries"]],
                [str(first.resolve())],
            )

    def test_batch_skip_business_does_not_create_collection_directory(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "course.mp4").write_bytes(b"placeholder")
            output = root / "output"
            argv = [str(PIPELINE), str(input_dir), str(output), "--skip-business"]

            with mock.patch.object(sys, "argv", argv):
                with mock.patch.object(run_video_course_pipeline, "run_batch_child", return_value=0):
                    with contextlib.redirect_stdout(io.StringIO()):
                        self.assertEqual(run_video_course_pipeline.main(), 0)

            self.assertFalse((output / "业务讲义汇总").exists())
            report = json.loads((output / "批量处理报告.json").read_text(encoding="utf-8"))
            self.assertIsNone(report["business_lectures_dir"])

    def test_dry_run_routes_custom_models_to_every_consuming_stage(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            result = run_pipeline(
                str(video),
                str(root / "output"),
                "--dry-run",
                "--visual-model",
                "visual-sentinel",
                "--asr-model",
                "asr-sentinel",
                "--text-model",
                "text-sentinel",
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            records = [json.loads(line) for line in result.stdout.splitlines()]

            visual = stage(records, "visual-analysis")["command"]
            transcription = stage(records, "audio-transcription")["command"]
            evidence = stage(records, "evidence-markdown")["command"]
            learner = stage(records, "learner-markdown-and-audit")["command"]
            self.assertEqual(option_value(visual, "--model"), "visual-sentinel")
            self.assertEqual(option_value(transcription, "--model"), "asr-sentinel")
            self.assertEqual(option_value(evidence, "--visual-model"), "visual-sentinel")
            self.assertEqual(option_value(evidence, "--asr-model"), "asr-sentinel")
            self.assertEqual(option_value(learner, "--model"), "text-sentinel")
            self.assertEqual(
                records[-1]["models"],
                {
                    "visual": "visual-sentinel",
                    "asr": "asr-sentinel",
                    "text": "text-sentinel",
                },
            )
            self.assertEqual(
                records[-1]["api_protocols"],
                {
                    "visual": "dashscope-native-multimodal-v1",
                    "asr": "dashscope-native-asr-transcription-v1",
                    "text": "dashscope-native-text-v1",
                },
            )

    def test_env_file_marks_key_set_without_printing_secret(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            env_file = root / "credentials.env"
            secret = "sk-never-print-this"
            env_file.write_text(f"DASHSCOPE_API_KEY={secret}\n", encoding="utf-8")
            result = run_pipeline(
                str(video), str(root / "output"), "--dry-run", "--env-file", str(env_file)
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertNotIn(secret, result.stdout + result.stderr)
            summary = json.loads(result.stdout.splitlines()[-1])
            self.assertTrue(summary["configuration"]["api_key_set"])
            self.assertTrue(summary["configuration"]["qwen_api_key_set"])
            self.assertTrue(summary["configuration"]["asr_api_key_set"])
            self.assertEqual(summary["configuration"]["env_source"], str(env_file.resolve()))

    def test_cli_unified_api_key_configures_all_models_without_leaking_or_propagating_argument(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "course.mp4").write_bytes(b"placeholder")
            secret = "sk-one-key-for-all-models"

            result = run_pipeline(
                str(input_dir),
                str(root / "output"),
                "--dry-run",
                "--api-key",
                secret,
            )

            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertNotIn(secret, result.stdout + result.stderr)
            records = [json.loads(line) for line in result.stdout.splitlines()]
            summary = records[-1]
            self.assertEqual(summary["configuration_source"], "cli --api-key")
            child_commands = [
                item["command"] for item in records if item.get("stage") == "adaptive-keyframes"
            ]
            self.assertTrue(child_commands)
            self.assertTrue(all(secret not in command for command in child_commands))

    def test_dedicated_keys_are_reported_separately_without_leaking_values(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            env_file = root / "credentials.env"
            qwen_secret = "sk-qwen-never-print"
            asr_secret = "sk-asr-never-print"
            env_file.write_text(
                f"DASHSCOPE_QWEN_API_KEY={qwen_secret}\n"
                f"DASHSCOPE_ASR_API_KEY={asr_secret}\n",
                encoding="utf-8",
            )
            result = run_pipeline(
                str(video), str(root / "output"), "--dry-run", "--env-file", str(env_file)
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertNotIn(qwen_secret, result.stdout + result.stderr)
            self.assertNotIn(asr_secret, result.stdout + result.stderr)
            configuration = json.loads(result.stdout.splitlines()[-1])["configuration"]
            self.assertTrue(configuration["qwen_api_key_set"])
            self.assertTrue(configuration["asr_api_key_set"])

    def test_real_run_requires_asr_key_when_asr_is_enabled(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            output = root / "output"
            env_file = root / "qwen-only.env"
            env_file.write_text("DASHSCOPE_QWEN_API_KEY=qwen-only\n", encoding="utf-8")
            result = run_pipeline(str(video), str(output), "--env-file", str(env_file))
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("DASHSCOPE_ASR_API_KEY", result.stderr)
            self.assertFalse(output.exists())

    def test_skip_asr_accepts_qwen_only_configuration(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            env_file = root / "qwen-only.env"
            env_file.write_text("DASHSCOPE_QWEN_API_KEY=qwen-only\n", encoding="utf-8")
            argv = [
                str(PIPELINE),
                str(video),
                str(root / "output"),
                "--env-file",
                str(env_file),
                "--skip-asr",
            ]
            with mock.patch.dict(os.environ, {}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(
                        run_video_course_pipeline,
                        "run_stage",
                        side_effect=RuntimeError("stage reached"),
                    ):
                        with self.assertRaisesRegex(RuntimeError, "stage reached"):
                            run_video_course_pipeline.main()

    def test_unsupported_video_extension_fails_before_any_stage(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "notes.txt"
            video.write_text("not video", encoding="utf-8")
            result = run_pipeline(str(video), str(root / "output"), "--dry-run")
            self.assertNotEqual(result.returncode, 0)
            error = json.loads(result.stderr)
            self.assertEqual(error["status"], "failed")
            self.assertIn("Unsupported video extension", error["error"])
            self.assertNotIn('"stage"', result.stdout)

    def test_output_path_that_is_a_file_is_rejected(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            output = root / "already-a-file"
            output.write_text("x", encoding="utf-8")
            result = run_pipeline(str(video), str(output), "--dry-run")
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("Output path is a file", result.stderr)

    def test_real_run_without_key_fails_before_creating_output(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            output = root / "output"
            env_file = root / "empty.env"
            env_file.write_text("# deliberately no DASHSCOPE_API_KEY\n", encoding="utf-8")
            result = run_pipeline(str(video), str(output), "--env-file", str(env_file))
            self.assertNotEqual(result.returncode, 0)
            self.assertIn("DASHSCOPE_API_KEY", result.stderr)
            self.assertFalse(output.exists())

    def test_batch_without_key_fails_before_starting_any_video(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            input_dir = root / "input"
            input_dir.mkdir()
            (input_dir / "a.mp4").write_bytes(b"a")
            (input_dir / "b.mp4").write_bytes(b"b")
            output = root / "output"
            empty_env = root / "empty.env"
            empty_env.write_text("# no key\n", encoding="utf-8")
            argv = [
                str(PIPELINE),
                str(input_dir),
                str(output),
                "--env-file",
                str(empty_env),
            ]

            with mock.patch.dict(os.environ, {}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(
                        run_video_course_pipeline, "run_batch_child"
                    ) as child:
                        with self.assertRaisesRegex(RuntimeError, "DASHSCOPE_API_KEY"):
                            run_video_course_pipeline.main()

            child.assert_not_called()
            self.assertFalse(output.exists())

    def test_single_model_failure_writes_structured_failure_report(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"video")
            output = root / "output"
            detail = {
                "event": "dashscope_error",
                "service": "visual",
                "category": "authentication",
                "http_status": 401,
                "provider_code": "InvalidApiKey",
                "message": "DashScope API Key 无效或已过期",
                "retryable": False,
            }

            def fail_visual(name: str, command: list[str], report: dict, dry_run: bool):
                report["stages"].append({"name": name, "command": command, "status": "complete"})
                if name == "adaptive-keyframes":
                    frames_dir = pathlib.Path(command[3])
                    frames_dir.mkdir(parents=True, exist_ok=True)
                    (frames_dir / "manifest.json").write_text(
                        json.dumps({"mode": "slides"}), encoding="utf-8"
                    )
                if name == "visual-analysis":
                    report["stages"][-1]["status"] = "failed"
                    raise run_video_course_pipeline.StageExecutionError(
                        1, command, error_detail=detail
                    )

            argv = [str(PIPELINE), str(video), str(output)]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(
                        run_video_course_pipeline, "run_stage", side_effect=fail_visual
                    ):
                        with contextlib.redirect_stdout(io.StringIO()):
                            with contextlib.redirect_stderr(io.StringIO()):
                                self.assertEqual(run_video_course_pipeline.cli(), 1)

            report_path = output / "course" / "course_处理报告.json"
            report = json.loads(report_path.read_text(encoding="utf-8"))
            self.assertEqual(report["status"], "failed")
            self.assertEqual(report["error_category"], "authentication")
            self.assertEqual(report["http_status"], 401)
            self.assertFalse(report["retryable"])

    def test_batch_stops_after_structured_authentication_failure(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            input_dir = root / "input"
            input_dir.mkdir()
            first = input_dir / "a.mp4"
            second = input_dir / "b.mp4"
            first.write_bytes(b"a")
            second.write_bytes(b"b")
            output = root / "output"

            def authentication_failure(command: list[str]) -> int:
                video = pathlib.Path(command[2])
                video_output = pathlib.Path(command[3])
                video_output.mkdir(parents=True, exist_ok=True)
                (video_output / f"{video.stem}_处理报告.json").write_text(
                    json.dumps(
                        {
                            "video": str(video),
                            "output_dir": str(video_output),
                            "status": "failed",
                            "error": "DashScope API Key 无效或已过期",
                            "error_category": "authentication",
                            "http_status": 401,
                            "retryable": False,
                            "models": {},
                            "stages": [],
                            "timing": {},
                        },
                        ensure_ascii=False,
                    ),
                    encoding="utf-8",
                )
                return 1

            argv = [str(PIPELINE), str(input_dir), str(output)]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(
                        run_video_course_pipeline,
                        "run_batch_child",
                        side_effect=authentication_failure,
                    ) as child:
                        with contextlib.redirect_stdout(io.StringIO()):
                            self.assertEqual(run_video_course_pipeline.main(), 1)

            self.assertEqual(child.call_count, 1)
            report = json.loads((output / "批量处理报告.json").read_text(encoding="utf-8"))
            self.assertTrue(report["stopped_early"])
            self.assertEqual(report["stop_category"], "authentication")
            self.assertEqual([item["status"] for item in report["results"]], ["failed", "skipped"])
            self.assertIn("API Key", report["results"][0]["error"])
            self.assertEqual(report["skipped_videos"], 1)

    def test_real_run_persists_models_and_configuration_in_report(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            output = root / "output"

            def complete_stage(name: str, command: list[str], report: dict, dry_run: bool) -> None:
                self.assertFalse(dry_run)
                report["stages"].append({"name": name, "command": command, "status": "complete"})
                if name == "adaptive-keyframes":
                    frames_dir = pathlib.Path(command[3])
                    frames_dir.mkdir(parents=True, exist_ok=True)
                    (frames_dir / "manifest.json").write_text(
                        json.dumps({"mode": "slides"}), encoding="utf-8"
                    )
                elif name == "visual-analysis":
                    visual_dir = pathlib.Path(command[3])
                    visual_dir.mkdir(parents=True, exist_ok=True)
                    (visual_dir / "consolidated.json").write_text(
                        json.dumps({"model": "visual-report", "groups": []}),
                        encoding="utf-8",
                    )

            stdout = io.StringIO()
            argv = [
                str(PIPELINE),
                str(video),
                str(output),
                "--visual-model",
                "visual-report",
                "--asr-model",
                "asr-report",
                "--text-model",
                "text-report",
                "--json-output",
            ]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(run_video_course_pipeline, "run_stage", complete_stage):
                        with contextlib.redirect_stdout(stdout):
                            self.assertEqual(run_video_course_pipeline.main(), 0)

            summary = json.loads(stdout.getvalue().splitlines()[-1])
            report_path = pathlib.Path(summary["report"])
            persisted = json.loads(report_path.read_text(encoding="utf-8"))
            self.assertEqual(
                persisted["models"],
                {"visual": "visual-report", "asr": "asr-report", "text": "text-report"},
            )
            self.assertEqual(
                persisted["configuration"],
                {
                    "env_source": "environment",
                    "api_key_set": True,
                    "qwen_api_key_set": True,
                    "asr_api_key_set": True,
                },
            )
            self.assertEqual(persisted["status"], "complete")
            log_path = pathlib.Path(summary["outputs"]["log_file"])
            log_records = [
                json.loads(line) for line in log_path.read_text(encoding="utf-8").splitlines()
            ]
            self.assertEqual(log_path.name, "视频处理日志.jsonl")
            self.assertEqual(log_records[0]["event"], "pipeline_started")
            self.assertEqual(log_records[-1]["event"], "pipeline_finished")
            self.assertNotIn("test-only", log_path.read_text(encoding="utf-8"))

    def test_real_run_writes_processing_stats_xlsx_and_report_usage(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            output = root / "output"
            elapsed = {
                "adaptive-keyframes": 1.0,
                "contact-sheet": 0.5,
                "visual-analysis": 2.0,
                "audio-transcription": 3.0,
                "evidence-markdown": 0.25,
                "learner-markdown-and-audit": 4.0,
            }

            def complete_stage(name: str, command: list[str], report: dict, dry_run: bool) -> None:
                report["stages"].append(
                    {
                        "name": name,
                        "command": command,
                        "status": "complete",
                        "elapsed_seconds": elapsed[name],
                    }
                )
                if name == "adaptive-keyframes":
                    frames_dir = pathlib.Path(command[3])
                    frames_dir.mkdir(parents=True, exist_ok=True)
                    (frames_dir / "manifest.json").write_text(
                        json.dumps({"mode": "slides"}), encoding="utf-8"
                    )
                elif name == "visual-analysis":
                    visual_dir = pathlib.Path(command[3])
                    visual_dir.mkdir(parents=True, exist_ok=True)
                    (visual_dir / "consolidated.json").write_text(
                        json.dumps(
                            {
                                "model": "qwen3-vl-flash",
                                "groups": [
                                    {
                                        "usage": {
                                            "prompt_tokens": 10,
                                            "completion_tokens": 5,
                                            "total_tokens": 15,
                                        }
                                    },
                                    {
                                        "usage": {
                                            "prompt_tokens": 100,
                                            "completion_tokens": 50,
                                            "total_tokens": 150,
                                        },
                                        "resumed": True,
                                    },
                                ],
                            }
                        ),
                        encoding="utf-8",
                    )
                elif name == "audio-transcription":
                    asr_dir = pathlib.Path(command[3])
                    asr_dir.mkdir(parents=True, exist_ok=True)
                    (asr_dir / "transcription.json").write_text(
                        json.dumps(
                            {
                                "transcripts": [
                                    {
                                        "content_duration_in_milliseconds": 12_500,
                                        "sentences": [],
                                    }
                                ]
                            }
                        ),
                        encoding="utf-8",
                    )
                elif name == "learner-markdown-and-audit":
                    audit_path = pathlib.Path(option_value(command, "--audit-output"))
                    audit_path.write_text(
                        json.dumps(
                            {
                                "usage": {
                                    "initial": {
                                        "prompt_tokens": 20,
                                        "completion_tokens": 6,
                                        "total_tokens": 26,
                                    },
                                    "repairs": [
                                        {
                                            "input_tokens": 3,
                                            "output_tokens": 1,
                                            "total_tokens": 4,
                                        }
                                    ],
                                }
                            }
                        ),
                        encoding="utf-8",
                    )

            stdout = io.StringIO()
            argv = [str(PIPELINE), str(video), str(output), "--json-output"]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(run_video_course_pipeline, "run_stage", complete_stage):
                        with contextlib.redirect_stdout(stdout):
                            self.assertEqual(run_video_course_pipeline.main(), 0)

            summary = json.loads(stdout.getvalue().splitlines()[-1])
            stats_path = pathlib.Path(summary["outputs"]["processing_stats"])
            self.assertTrue(stats_path.is_file())
            workbook = load_workbook(stats_path, data_only=True)
            detail = workbook["逐视频统计"]
            headers = [cell.value for cell in detail[1]]
            rows = [dict(zip(headers, values)) for values in detail.iter_rows(min_row=2, values_only=True)]
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["视频路径"], str(video.resolve()))
            self.assertEqual(rows[0]["视觉总Token"], 15)
            self.assertEqual(rows[0]["文本总Token"], 30)
            self.assertEqual(rows[0]["单视频总Token"], 45)
            self.assertEqual(rows[0]["ASR音频时长（秒）"], 12.5)
            self.assertEqual(rows[0]["视觉分析耗时（秒）"], 2.0)
            batch_sheet = workbook["批次汇总"]
            batch_values = dict(
                zip(
                    [cell.value for cell in batch_sheet[1]],
                    [cell.value for cell in batch_sheet[2]],
                )
            )
            self.assertEqual(batch_values["批次总Token"], 45)
            workbook.close()

            report = json.loads(pathlib.Path(summary["report"]).read_text(encoding="utf-8"))
            self.assertEqual(report["usage"]["qwen_total_tokens"], 45)
            self.assertEqual(report["usage"]["asr_audio_seconds"], 12.5)
            self.assertIn("started_at", report["timing"])
            self.assertIn("finished_at", report["timing"])

    def test_single_run_groups_outputs_into_video_subdirectory(self):
        r"""单跑模式下，所有产物应自动进 <输出目录>\<标题>\ 子目录，
        与批处理模式路径结构保持一致，避免散落到根目录产生孤儿目录。
        """
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            output_root = root / "output"

            def complete_stage(name: str, command: list[str], report: dict, dry_run: bool) -> None:
                self.assertFalse(dry_run)
                report["stages"].append({"name": name, "command": command, "status": "complete"})
                if name == "adaptive-keyframes":
                    frames_dir = pathlib.Path(command[3])
                    frames_dir.mkdir(parents=True, exist_ok=True)
                    (frames_dir / "manifest.json").write_text(
                        json.dumps({"mode": "slides"}), encoding="utf-8"
                    )
                elif name == "visual-analysis":
                    visual_dir = pathlib.Path(command[3])
                    visual_dir.mkdir(parents=True, exist_ok=True)
                    (visual_dir / "consolidated.json").write_text(
                        json.dumps({"model": "qwen3-vl-flash", "groups": []}),
                        encoding="utf-8",
                    )
                elif name == "audio-transcription":
                    asr_dir = pathlib.Path(command[3])
                    asr_dir.mkdir(parents=True, exist_ok=True)
                    (asr_dir / "transcription.json").write_text(
                        json.dumps({"transcripts": []}), encoding="utf-8"
                    )

            argv = [str(PIPELINE), str(video), str(output_root)]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(run_video_course_pipeline, "run_stage", complete_stage):
                        with contextlib.redirect_stdout(io.StringIO()):
                            self.assertEqual(run_video_course_pipeline.main(), 0)

            # 关键断言：产物不在 output_root 根下，而在 output_root / "course" 子目录下
            self.assertFalse((output_root / "视频处理统计.xlsx").exists())
            self.assertFalse((output_root / "_video_course_work").exists())
            self.assertFalse((output_root / "course_处理报告.json").exists())

            self.assertTrue((output_root / "course").is_dir())
            self.assertTrue((output_root / "course" / "视频处理统计.xlsx").is_file())
            self.assertTrue((output_root / "course" / "course_处理报告.json").is_file())
            self.assertTrue((output_root / "course" / "_video_course_work" / "course").is_dir())

    def test_report_and_evidence_stage_use_visual_artifact_producer_model(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            video = root / "course.mp4"
            video.write_bytes(b"placeholder")
            output = root / "output"

            def complete_stage(name: str, command: list[str], report: dict, dry_run: bool) -> None:
                report["stages"].append({"name": name, "command": command, "status": "complete"})
                if name == "adaptive-keyframes":
                    frames_dir = pathlib.Path(command[3])
                    frames_dir.mkdir(parents=True, exist_ok=True)
                    (frames_dir / "manifest.json").write_text(
                        json.dumps({"mode": "slides"}), encoding="utf-8"
                    )
                elif name == "visual-analysis":
                    visual_dir = pathlib.Path(command[3])
                    visual_dir.mkdir(parents=True, exist_ok=True)
                    (visual_dir / "consolidated.json").write_text(
                        json.dumps({"model": "producer-resolved-visual", "groups": []}),
                        encoding="utf-8",
                    )

            stdout = io.StringIO()
            argv = [
                str(PIPELINE),
                str(video),
                str(output),
                "--visual-model",
                "requested-visual-alias",
                "--skip-asr",
                "--skip-business",
                "--json-output",
            ]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(run_video_course_pipeline, "run_stage", complete_stage):
                        with contextlib.redirect_stdout(stdout):
                            self.assertEqual(run_video_course_pipeline.main(), 0)

            summary = json.loads(stdout.getvalue().splitlines()[-1])
            report = json.loads(pathlib.Path(summary["report"]).read_text(encoding="utf-8"))
            evidence_command = next(
                item["command"] for item in report["stages"] if item["name"] == "evidence-markdown"
            )
            self.assertEqual(report["models"]["visual"], "producer-resolved-visual")
            self.assertIsNone(report["models"]["asr"])
            self.assertEqual(
                option_value(evidence_command, "--visual-model"),
                "producer-resolved-visual",
            )
            self.assertEqual(
                option_value(evidence_command, "--asr-model"),
                "未执行（--skip-asr）",
            )


class StageDefaultTests(unittest.TestCase):
    def test_visual_analysis_default_model_is_written_to_consolidated_output(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            frame = root / "frame.jpg"
            frame.write_bytes(b"image")
            video = root / "course.mp4"
            video.write_bytes(b"video")
            manifest = root / "manifest.json"
            manifest.write_text(
                json.dumps(
                    {
                        "video": str(video),
                        "keyframes": [{"timestamp_seconds": 0, "image": str(frame)}],
                    }
                ),
                encoding="utf-8",
            )
            output = root / "visual"
            argv = ["analyze_video_frames.py", str(manifest), str(output)]
            group_result = {
                "group_id": 1,
                "source_frames": [{"timestamp_seconds": 0, "image": str(frame)}],
                "result": {"frames": []},
                "usage": {},
                "model": "qwen3-vl-flash",
            }
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(analyze_video_frames, "invoke_group", return_value=group_result):
                        with contextlib.redirect_stdout(io.StringIO()):
                            self.assertEqual(analyze_video_frames.main(), 0)
            consolidated = json.loads((output / "consolidated.json").read_text(encoding="utf-8"))
            self.assertEqual(consolidated["model"], "qwen3-vl-flash")

    def test_learner_rewrite_default_model_is_sent_to_model_call(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            source = root / "evidence.md"
            source.write_text("# Course\n", encoding="utf-8")
            output = root / "learner.md"
            argv = ["rewrite_learner_markdown.py", str(source), str(output)]
            with mock.patch.dict(os.environ, {"DASHSCOPE_API_KEY": "test-only"}, clear=True):
                with mock.patch.object(sys, "argv", argv):
                    with mock.patch.object(
                        rewrite_learner_markdown,
                        "call_model",
                        return_value=("# Learner", {}),
                    ) as call_model:
                        with contextlib.redirect_stdout(io.StringIO()):
                            self.assertEqual(rewrite_learner_markdown.main(), 0)
            self.assertEqual(call_model.call_args.args[1], "qwen-plus")

    def test_evidence_assembly_defaults_and_visual_json_fallback_are_observable(self):
        with tempfile.TemporaryDirectory() as raw:
            root = pathlib.Path(raw)
            visual = root / "visual.json"
            visual.write_text(json.dumps({"video": "course.mp4", "frames": []}), encoding="utf-8")
            asr = root / "asr.json"
            asr.write_text(json.dumps({"transcripts": []}), encoding="utf-8")
            audio = root / "audio.mp3"
            audio.write_bytes(b"audio")
            markdown = root / "evidence.md"
            visual_output = root / "visual-output.json"
            argv = [
                "assemble_evidence_markdown.py",
                "--title",
                "Course",
                "--mode",
                "slides",
                "--visual",
                str(visual),
                "--asr",
                str(asr),
                "--output",
                str(markdown),
                "--assets",
                str(root / "assets"),
                "--visual-json-output",
                str(visual_output),
                "--asr-json-output",
                str(root / "asr-output.json"),
                "--audio-source",
                str(audio),
                "--audio-output",
                str(root / "audio-output.mp3"),
            ]
            with mock.patch.object(sys, "argv", argv):
                with contextlib.redirect_stdout(io.StringIO()):
                    self.assertEqual(assemble_evidence_markdown.main(), 0)
            self.assertIn("qwen3-vl-flash", markdown.read_text(encoding="utf-8"))
            self.assertIn("paraformer-v2", markdown.read_text(encoding="utf-8"))
            self.assertEqual(
                json.loads(visual_output.read_text(encoding="utf-8"))["model"],
                "qwen3-vl-flash",
            )


class SkillDocumentationTests(unittest.TestCase):
    def test_skill_documents_local_contract_models_and_env(self):
        skill = (ROOT / "SKILL.md").read_text(encoding="utf-8")
        config = (ROOT / "references" / "configuration.md").read_text(encoding="utf-8")
        metadata = (ROOT / "agents" / "openai.yaml").read_text(encoding="utf-8")
        combined = skill + config
        for expected in (
            "qwen3-vl-flash",
            "paraformer-v2",
            "qwen-plus",
            "--env-file",
            "DASHSCOPE_QWEN_API_KEY",
            "DASHSCOPE_ASR_API_KEY",
            "DASHSCOPE_API_KEY",
            "视频处理统计.xlsx",
            "批量处理报告.json",
            "--recursive",
            "文件夹路径",
            "单视频总Token",
            "批次总Token",
            "ASR音频时长",
            "本地",
            "DashScope",
        ):
            self.assertIn(expected, combined)
        self.assertIn("DashScope 原生", combined)
        self.assertIn("multimodal-generation/generation", combined)
        self.assertIn("text-generation/generation", combined)
        self.assertNotIn("使用兼容 OpenAI 格式的文本接口", combined)
        self.assertIn("$video-course-to-markdown", metadata)

    def test_skill_documents_exact_configuration_and_local_cli_contract(self):
        skill = (ROOT / "SKILL.md").read_text(encoding="utf-8")
        config = (ROOT / "references" / "configuration.md").read_text(encoding="utf-8")
        combined = skill + config
        self.assertIn("MP4、MOV、MKV、AVI、WEBM 或 M4V", combined)
        self.assertIn("用户指定的本地目录", combined)
        self.assertIn('"D:\\courses\\service-training.mp4"', skill)
        self.assertIn('"D:\\courses\\service-training-output"', skill)
        self.assertIn("--mode auto", skill)
        self.assertIn('--env-file "D:\\secure\\video-course.env"', skill)
        self.assertIn('"D:\\courses\\incoming"', skill)
        self.assertIn('"D:\\courses\\batch-output"', skill)
        self.assertIn("关键帧、音频和文本会发送到阿里云百炼 DashScope 处理", combined)
        self.assertIn("Markdown、JSON、日志或 Skill 包", combined)
        self.assertIn("Existing process environment variables", config)
        self.assertIn("提供 --env-file 时，只读取该文件", config)
        self.assertIn("不查找项目或用户私有 .env", config)
        self.assertIn("未提供 --env-file", config)
        self.assertIn("<skill-root>/.env", config)
        self.assertIn("%LOCALAPPDATA%\\Codex\\video-course-to-markdown\\.env", config)
        self.assertIn("`DASHSCOPE_QWEN_API_KEY` → `DASHSCOPE_API_KEY`", config)
        self.assertIn("`DASHSCOPE_ASR_API_KEY` → `DASHSCOPE_API_KEY`", config)
        self.assertIn("`--skip-asr` 时不要求 ASR Key", config)
        self.assertIn("断点续跑复用的旧视觉结果不重复计入本次 Token", config)
        for row in (
            "| 画面文字和视觉理解 | `qwen3-vl-flash` | `--visual-model` |",
            "| 带时间戳音频转写 | `paraformer-v2` | `--asr-model` |",
            "| 学员稿转换与缺失修复 | `qwen-plus` | `--text-model` |",
        ):
            self.assertIn(row, config)

    def test_agent_metadata_is_safe_and_invokes_the_skill(self):
        metadata = yaml.safe_load((ROOT / "agents" / "openai.yaml").read_text(encoding="utf-8"))
        self.assertEqual(set(metadata), {"interface"})
        interface = metadata["interface"]
        self.assertEqual(set(interface), {"display_name", "short_description", "default_prompt"})
        self.assertGreaterEqual(len(interface["short_description"]), 25)
        self.assertLessEqual(len(interface["short_description"]), 64)
        self.assertIn("$video-course-to-markdown", interface["default_prompt"])
        self.assertNotIn("mcp", str(metadata).lower())
        self.assertNotIn("DASHSCOPE_API_KEY", str(metadata))

    def test_real_secret_is_not_committed(self):
        example = (ROOT / ".env.example").read_text(encoding="utf-8")
        self.assertIn("DASHSCOPE_API_KEY=sk-your-unified-api-key", example)
        self.assertIn("# DASHSCOPE_QWEN_API_KEY=sk-your-qwen-api-key", example)
        self.assertIn("# DASHSCOPE_ASR_API_KEY=sk-your-asr-api-key", example)
        self.assertNotIn("sk-real", example)
        self.assertIn(".env", (ROOT / ".gitignore").read_text(encoding="utf-8").splitlines())
