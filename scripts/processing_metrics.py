import datetime
import os
import pathlib
import time
import unicodedata
from collections.abc import Iterable


DETAIL_SHEET = "逐视频统计"
MODEL_SUMMARY_SHEET = "模型Token汇总"
BATCH_SUMMARY_SHEET = "批次汇总"

XLSX_HEADERS = [
    "视频路径", "状态", "开始时间", "结束时间", "总耗时（秒）",
    "抽帧耗时（秒）", "关键帧总览耗时（秒）", "视觉分析耗时（秒）",
    "音频转写耗时（秒）", "证据稿耗时（秒）", "业务讲义耗时（秒）",
    "视觉模型", "视觉输入Token", "视觉输出Token", "视觉总Token",
    "ASR模型", "ASR Token计量", "ASR音频时长（秒）",
    "文本模型", "文本输入Token", "文本输出Token", "文本总Token",
    "单视频总Token", "处理报告", "错误信息",
]

MODEL_SUMMARY_HEADERS = [
    "模型类型", "模型名称", "输入Token", "输出Token", "总Token",
    "ASR音频时长（秒）", "计量说明",
]

BATCH_SUMMARY_HEADERS = [
    "统计更新时间", "视频总数", "成功视频数", "失败视频数", "其他状态视频数",
    "视频处理总耗时（秒）", "视觉总Token", "文本总Token", "批次总Token",
    "ASR音频总时长（秒）",
]


def _nonnegative_int(value: object) -> int:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return 0
    return max(0, int(value))


def _nonnegative_number(value: object) -> float:
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return max(0.0, float(value))
    if isinstance(value, str):
        try:
            return max(0.0, float(value))
        except ValueError:
            return 0.0
    return 0.0


def normalize_token_usage(usage: object) -> dict[str, int]:
    if not isinstance(usage, dict):
        usage = {}
    input_tokens = _nonnegative_int(
        usage.get("prompt_tokens", usage.get("input_tokens", 0))
    )
    output_tokens = _nonnegative_int(
        usage.get("completion_tokens", usage.get("output_tokens", 0))
    )
    total_value = usage.get("total_tokens")
    total_tokens = (
        _nonnegative_int(total_value)
        if total_value is not None
        else input_tokens + output_tokens
    )
    return {
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": total_tokens,
    }


def _sum_token_usage(usages: Iterable[object]) -> dict[str, int]:
    total = {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0}
    for usage in usages:
        normalized = normalize_token_usage(usage)
        for key in total:
            total[key] += normalized[key]
    return total


def summarize_token_usage(
    visual_data: object,
    audit_data: object,
) -> dict[str, object]:
    visual_groups = visual_data.get("groups", []) if isinstance(visual_data, dict) else []
    visual_usages = [
        group.get("usage", {})
        for group in visual_groups
        if isinstance(group, dict) and not group.get("resumed")
    ]

    text_usages: list[object] = []
    if isinstance(audit_data, dict) and isinstance(audit_data.get("usage"), dict):
        usage = audit_data["usage"]
        text_usages.append(usage.get("initial", {}))
        repairs = usage.get("repairs", [])
        if isinstance(repairs, list):
            text_usages.extend(repairs)

    visual = _sum_token_usage(visual_usages)
    text = _sum_token_usage(text_usages)
    total_tokens = visual["total_tokens"] + text["total_tokens"]
    return {
        "visual": visual,
        "text": text,
        "qwen_total_tokens": total_tokens,
        "total_tokens": total_tokens,
    }


def transcription_duration_seconds(transcription: object) -> float:
    if not isinstance(transcription, dict):
        return 0.0
    duration_ms = 0
    transcripts = transcription.get("transcripts", [])
    if not isinstance(transcripts, list):
        return 0.0
    for transcript in transcripts:
        if not isinstance(transcript, dict):
            continue
        duration_ms = max(
            duration_ms,
            _nonnegative_int(transcript.get("content_duration_in_milliseconds", 0)),
        )
        sentences = transcript.get("sentences", [])
        if not isinstance(sentences, list):
            continue
        for sentence in sentences:
            if isinstance(sentence, dict):
                duration_ms = max(duration_ms, _nonnegative_int(sentence.get("end_time", 0)))
    return round(duration_ms / 1000, 3)


def _stage_elapsed(report: dict, name: str) -> float:
    elapsed = 0.0
    for stage in report.get("stages", []):
        if isinstance(stage, dict) and stage.get("name") == name:
            value = stage.get("elapsed_seconds", 0)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                elapsed += max(0.0, float(value))
    return round(elapsed, 2)


def build_processing_row(
    report: dict,
    usage: dict,
    *,
    asr_duration_seconds: float,
    started_at: str,
    finished_at: str,
    total_elapsed_seconds: float,
    report_path: pathlib.Path,
) -> dict[str, object]:
    visual = usage.get("visual", {})
    text = usage.get("text", {})
    models = report.get("models", {})
    total_tokens = usage.get("total_tokens", usage.get("qwen_total_tokens", 0))
    asr_model = models.get("asr") or ""
    return {
        "视频路径": report.get("video", ""),
        "状态": report.get("status", ""),
        "开始时间": started_at,
        "结束时间": finished_at,
        "总耗时（秒）": round(max(0.0, total_elapsed_seconds), 2),
        "抽帧耗时（秒）": _stage_elapsed(report, "adaptive-keyframes"),
        "关键帧总览耗时（秒）": _stage_elapsed(report, "contact-sheet"),
        "视觉分析耗时（秒）": _stage_elapsed(report, "visual-analysis"),
        "音频转写耗时（秒）": _stage_elapsed(report, "audio-transcription"),
        "证据稿耗时（秒）": _stage_elapsed(report, "evidence-markdown"),
        "业务讲义耗时（秒）": _stage_elapsed(report, "learner-markdown-and-audit"),
        "视觉模型": models.get("visual") or "",
        "视觉输入Token": visual.get("input_tokens", 0),
        "视觉输出Token": visual.get("output_tokens", 0),
        "视觉总Token": visual.get("total_tokens", 0),
        "ASR模型": asr_model,
        "ASR Token计量": "不适用（按音频时长计费）" if asr_model else "未执行",
        "ASR音频时长（秒）": round(max(0.0, asr_duration_seconds), 3),
        "文本模型": models.get("text") or "",
        "文本输入Token": text.get("input_tokens", 0),
        "文本输出Token": text.get("output_tokens", 0),
        "文本总Token": text.get("total_tokens", 0),
        "单视频总Token": total_tokens,
        "处理报告": str(report_path),
        "错误信息": "",
    }


def _replace_stats_file(temporary: pathlib.Path, path: pathlib.Path) -> None:
    """Atomically replace the XLSX workbook, tolerating brief Windows locks."""
    last_error: PermissionError | None = None
    for attempt in range(5):
        try:
            os.replace(temporary, path)
            return
        except PermissionError as error:
            last_error = error
            time.sleep(0.2 * (attempt + 1))
    raise PermissionError(
        f"无法更新统计工作簿 {path}（{last_error}）。请关闭 Excel 或其他占用该文件的程序后重跑。"
    ) from last_error


def _load_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError(
            "Writing 视频处理统计.xlsx requires openpyxl. "
            "Run scripts/bootstrap_dependencies.py and pass its target with --pydeps."
        ) from error
    return openpyxl, Alignment, Font, PatternFill, get_column_letter


def validate_xlsx_support() -> None:
    """Fail before expensive video stages when XLSX support is unavailable."""
    _load_openpyxl()


def _read_existing_rows(path: pathlib.Path) -> list[dict[str, object]]:
    if not path.is_file():
        return []
    openpyxl, _, _, _, _ = _load_openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if DETAIL_SHEET not in workbook.sheetnames:
            return []
        sheet = workbook[DETAIL_SHEET]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1), ())
        headers = [cell.value for cell in header_row]
        rows = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            existing = {str(header): value for header, value in zip(headers, values) if header}
            if existing.get("视频路径"):
                if "单视频总Token" not in existing and "Qwen总Token" in existing:
                    existing["单视频总Token"] = existing["Qwen总Token"]
                rows.append(existing)
        return rows
    finally:
        workbook.close()


def _style_sheet(sheet, headers: list[str], *, freeze: str = "A2") -> None:
    _, Alignment, Font, PatternFill, get_column_letter = _load_openpyxl()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 30
    def display_width(value: str) -> int:
        return sum(
            2 if unicodedata.east_asian_width(character) in {"W", "F", "A"} else 1
            for character in value
        )

    for index, header in enumerate(headers, start=1):
        values = [header]
        for row in sheet.iter_rows(min_row=2, min_col=index, max_col=index, values_only=True):
            if row[0] is not None:
                values.append(str(row[0]))
        width = min(52, max(12, max(display_width(value) for value in values) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, header in enumerate(headers, start=1):
        column = sheet.get_column_letter(index) if hasattr(sheet, "get_column_letter") else None
        if column is None:
            column = get_column_letter(index)
        if "Token" in header and "计量" not in header:
            for cell in sheet[column][1:]:
                cell.number_format = "#,##0"
        elif "（秒）" in header:
            for cell in sheet[column][1:]:
                cell.number_format = "0.000"


def _model_summary_rows(rows: list[dict[str, object]]) -> list[list[object]]:
    aggregates: dict[tuple[str, str], dict[str, float]] = {}
    for row in rows:
        for model_type, model_field, prefix in (
            ("视觉", "视觉模型", "视觉"),
            ("文本", "文本模型", "文本"),
        ):
            model = str(row.get(model_field) or "").strip()
            if not model:
                continue
            aggregate = aggregates.setdefault(
                (model_type, model),
                {"input": 0.0, "output": 0.0, "total": 0.0, "audio": 0.0},
            )
            aggregate["input"] += _nonnegative_number(row.get(f"{prefix}输入Token"))
            aggregate["output"] += _nonnegative_number(row.get(f"{prefix}输出Token"))
            aggregate["total"] += _nonnegative_number(row.get(f"{prefix}总Token"))

        asr_model = str(row.get("ASR模型") or "").strip()
        if asr_model:
            aggregate = aggregates.setdefault(
                ("ASR", asr_model),
                {"input": 0.0, "output": 0.0, "total": 0.0, "audio": 0.0},
            )
            aggregate["audio"] += _nonnegative_number(row.get("ASR音频时长（秒）"))

    order = {"视觉": 0, "文本": 1, "ASR": 2}
    output = []
    for (model_type, model), aggregate in sorted(
        aggregates.items(), key=lambda item: (order[item[0][0]], item[0][1].casefold())
    ):
        is_asr = model_type == "ASR"
        output.append([
            model_type,
            model,
            None if is_asr else int(aggregate["input"]),
            None if is_asr else int(aggregate["output"]),
            None if is_asr else int(aggregate["total"]),
            round(aggregate["audio"], 3) if is_asr else None,
            "不提供Token，按音频时长计费" if is_asr else "按Token统计",
        ])
    return output


def _batch_summary_row(rows: list[dict[str, object]]) -> list[object]:
    completed = sum(str(row.get("状态") or "") == "complete" for row in rows)
    failed = sum(str(row.get("状态") or "") == "failed" for row in rows)
    return [
        datetime.datetime.now().astimezone().isoformat(timespec="seconds"),
        len(rows),
        completed,
        failed,
        len(rows) - completed - failed,
        round(sum(_nonnegative_number(row.get("总耗时（秒）")) for row in rows), 2),
        int(sum(_nonnegative_number(row.get("视觉总Token")) for row in rows)),
        int(sum(_nonnegative_number(row.get("文本总Token")) for row in rows)),
        int(sum(_nonnegative_number(row.get("单视频总Token")) for row in rows)),
        round(sum(_nonnegative_number(row.get("ASR音频时长（秒）")) for row in rows), 3),
    ]


def _write_workbook(path: pathlib.Path, rows: list[dict[str, object]]) -> None:
    openpyxl, _, _, _, _ = _load_openpyxl()
    workbook = openpyxl.Workbook()
    detail = workbook.active
    detail.title = DETAIL_SHEET
    detail.append(XLSX_HEADERS)
    for row in rows:
        detail.append([row.get(header, "") for header in XLSX_HEADERS])
    _style_sheet(detail, XLSX_HEADERS, freeze="C2")

    models = workbook.create_sheet(MODEL_SUMMARY_SHEET)
    models.append(MODEL_SUMMARY_HEADERS)
    for model_row in _model_summary_rows(rows):
        models.append(model_row)
    _style_sheet(models, MODEL_SUMMARY_HEADERS)

    batch = workbook.create_sheet(BATCH_SUMMARY_SHEET)
    batch.append(BATCH_SUMMARY_HEADERS)
    batch.append(_batch_summary_row(rows))
    _style_sheet(batch, BATCH_SUMMARY_HEADERS)

    for sheet in (detail, models, batch):
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"

    temporary = path.with_name(f".{path.stem}.{os.getpid()}.tmp.xlsx")
    try:
        workbook.save(temporary)
        workbook.close()
        _replace_stats_file(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def upsert_processing_stats(path: pathlib.Path, row: dict[str, object]) -> None:
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Processing statistics must use an .xlsx path, got: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = [
        existing
        for existing in _read_existing_rows(path)
        if existing.get("视频路径") != row.get("视频路径")
    ]
    rows.append({header: row.get(header, "") for header in XLSX_HEADERS})
    _write_workbook(path, rows)
